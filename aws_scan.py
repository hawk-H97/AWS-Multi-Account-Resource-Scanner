#!/usr/bin/env python3
"""
aws_scan.py
───────────
Scans ALL AWS resources using direct boto3 API calls across ALL regions.
No Resource Explorer required.

Output columns (exact):
  Account ID | Resource Name | Resource Type | Service | ARN | Region |
  Owning Account | Tags | Created Date | Created By | Status | Unassigned |
  Last Reported At

Output file: aws_inventory_ACCOUNTID_YYYYMMDD_HHMM.xlsx
"""

import boto3
import datetime
import json
import os
import signal
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── EXACT columns as requested ────────────────────────────────────────────────
COLS = [
    "Account ID",
    "Resource Name",
    "Resource Type",
    "Service",
    "ARN",
    "Region",
    "Owning Account",
    "Tags",
    "Created Date",
    "Created By",
    "Status",
    "Unassigned",
    "Last Reported At",
]

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='0D2B55')
ALT_FILL     = PatternFill('solid', fgColor='EDF4FF')
RED_FILL     = PatternFill('solid', fgColor='FFCCCC')
ORANGE_FILL  = PatternFill('solid', fgColor='FFF3CD')
GREEN_FILL   = PatternFill('solid', fgColor='E8F5E9')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
BODY_FONT    = Font(name='Calibri', size=10)
TITLE_FONT   = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN   = Alignment(vertical='center', wrap_text=False)
THIN         = Side(style='thin', color='CCCCCC')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SVC_FILLS    = [
    PatternFill('solid', fgColor='DDEEFF'),
    PatternFill('solid', fgColor='E8F5E9'),
    PatternFill('solid', fgColor='FFF3CD'),
    PatternFill('solid', fgColor='FCE4EC'),
    PatternFill('solid', fgColor='F3E5F5'),
    PatternFill('solid', fgColor='E0F7FA'),
    PatternFill('solid', fgColor='FBE9E7'),
    PatternFill('solid', fgColor='F1F8E9'),
]

ACTIVE_WORDS = {
    'running','available','in-use','attached','associated','active',
    'healthy','enabled','create_complete','in_sync','issued','inservice',
    'backing-up','modifying','provisioning','update_complete','completed',
}
INACTIVE_WORDS = {
    'stopped','stopping','terminated','terminating','shutting-down',
    'deleted','deleting','error','unassociated','unattached','detached',
    'unused','unhealthy','draining','failed','inactive','disabled',
    'create_failed','rollback_complete','rollback_failed','never logged in',
    'expired','revoked','unavailable','alarm','outofservice',
    'inaccessible-encryption-credentials','restore-error',
}

# ── CloudTrail event sets ─────────────────────────────────────────────────────
CREATE_EVENTS = {
    'RunInstances','CreateVolume','CreateSnapshot','CreateSecurityGroup',
    'CreateVpc','CreateSubnet','CreateNatGateway','AllocateAddress',
    'CreateInternetGateway','CreateRouteTable','CreateNetworkInterface',
    'CreateLaunchTemplate','CreateKeyPair','CreateImage',
    'CreateBucket','CreateDBInstance','CreateDBCluster',
    'CreateFunction20150331','CreateRole','CreateUser','CreatePolicy',
    'CreateGroup','CreateCluster','CreateService','CreateTable',
    'CreateLoadBalancer','CreateStack','CreateSecret','CreateTopic',
    'CreateQueue','CreateKey','CreateLogGroup','CreateStateMachine',
    'CreateRepository','CreateCacheCluster','PutParameter',
    'CreateProject','CreatePipeline','CreateRestApi','CreateJob',
    'CreateDatabase','CreateDomain',
}

# ── Helpers ───────────────────────────────────────────────────────────────────
NOW_UTC = datetime.datetime.utcnow()


def get_account_id():
    return boto3.client('sts').get_caller_identity()['Account']


def get_all_regions():
    ec2 = boto3.client('ec2', region_name='us-east-1')
    return sorted([r['RegionName']
                   for r in ec2.describe_regions(AllRegions=False)['Regions']])


def fmt_date(val):
    if not val:
        return ''
    s = str(val)
    return s[:10]


def fmt_tags(tag_list):
    if not tag_list:
        return ''
    if isinstance(tag_list, list):
        return ', '.join(
            f"{t.get('Key','').strip()}={t.get('Value','').strip()}"
            for t in tag_list if isinstance(t, dict) and t.get('Key')
        )
    if isinstance(tag_list, dict):
        return ', '.join(f"{k}={v}" for k, v in tag_list.items())
    return str(tag_list)


def make_rec(account_id, name, rtype, service, arn, region,
             tags=None, created='', status='', unassigned=''):
    """Build a record with ALL required columns pre-filled."""
    return {
        "Account ID"      : account_id,
        "Resource Name"   : name or '',
        "Resource Type"   : rtype,
        "Service"         : service,
        "ARN"             : arn,
        "Region"          : region,
        "Owning Account"  : account_id,      # filled; overridden if cross-account
        "Tags"            : fmt_tags(tags) if isinstance(tags, list) else (tags or ''),
        "Created Date"    : created,
        "Created By"      : '',              # filled by CloudTrail pass
        "Status"          : status,
        "Unassigned"      : unassigned,
        "Last Reported At": NOW_UTC.strftime('%Y-%m-%d %H:%M UTC'),
    }

# ══════════════════════════════════════════════════════════════════════════════
# SERVICE SCANNERS
# ══════════════════════════════════════════════════════════════════════════════

def scan_ec2_instances(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_instances').paginate():
            for res in page['Reservations']:
                for i in res['Instances']:
                    iid  = i['InstanceId']
                    name = next((t['Value'] for t in i.get('Tags', [])
                                 if t['Key'] == 'Name'), iid)
                    arn  = f"arn:aws:ec2:{region}:{account_id}:instance/{iid}"
                    recs.append(make_rec(
                        account_id, name, 'ec2:instance', 'ec2', arn, region,
                        tags=i.get('Tags', []),
                        created=fmt_date(i.get('LaunchTime')),
                        status=i.get('State', {}).get('Name', ''),
                    ))
    except Exception:
        pass
    return recs


def scan_ec2_volumes(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_volumes').paginate():
            for v in page['Volumes']:
                vid  = v['VolumeId']
                name = next((t['Value'] for t in v.get('Tags', [])
                             if t['Key'] == 'Name'), vid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:volume/{vid}"
                atts = v.get('Attachments', [])
                recs.append(make_rec(
                    account_id, name, 'ec2:volume', 'ec2', arn, region,
                    tags=v.get('Tags', []),
                    created=fmt_date(v.get('CreateTime')),
                    status=v.get('State', ''),
                    unassigned='No' if atts else 'Yes',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_snapshots(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_snapshots').paginate(OwnerIds=[account_id]):
            for s in page['Snapshots']:
                sid  = s['SnapshotId']
                name = next((t['Value'] for t in s.get('Tags', [])
                             if t['Key'] == 'Name'), sid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:snapshot/{sid}"
                recs.append(make_rec(
                    account_id, name, 'ec2:snapshot', 'ec2', arn, region,
                    tags=s.get('Tags', []),
                    created=fmt_date(s.get('StartTime')),
                    status=s.get('State', ''),
                ))
    except Exception:
        pass
    return recs


def scan_ec2_security_groups(ec2, account_id, region):
    recs = []
    try:
        used = set()
        for page in ec2.get_paginator('describe_network_interfaces').paginate():
            for ni in page['NetworkInterfaces']:
                for sg in ni.get('Groups', []):
                    used.add(sg['GroupId'])
        for page in ec2.get_paginator('describe_security_groups').paginate():
            for sg in page['SecurityGroups']:
                sgid = sg['GroupId']
                arn  = f"arn:aws:ec2:{region}:{account_id}:security-group/{sgid}"
                recs.append(make_rec(
                    account_id, sg.get('GroupName', sgid),
                    'ec2:securityGroup', 'ec2', arn, region,
                    tags=sg.get('Tags', []),
                    status='in-use' if sgid in used else 'unused',
                    unassigned='No' if sgid in used else 'Yes',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_vpcs(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_vpcs').paginate():
            for v in page['Vpcs']:
                vid  = v['VpcId']
                name = next((t['Value'] for t in v.get('Tags', [])
                             if t['Key'] == 'Name'), vid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:vpc/{vid}"
                recs.append(make_rec(
                    account_id, name, 'ec2:vpc', 'ec2', arn, region,
                    tags=v.get('Tags', []),
                    status='default' if v.get('IsDefault') else 'available',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_subnets(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_subnets').paginate():
            for s in page['Subnets']:
                sid  = s['SubnetId']
                name = next((t['Value'] for t in s.get('Tags', [])
                             if t['Key'] == 'Name'), sid)
                arn  = s.get('SubnetArn', f"arn:aws:ec2:{region}:{account_id}:subnet/{sid}")
                recs.append(make_rec(
                    account_id, name, 'ec2:subnet', 'ec2', arn, region,
                    tags=s.get('Tags', []),
                    status=s.get('State', ''),
                ))
    except Exception:
        pass
    return recs


def scan_ec2_eips(ec2, account_id, region):
    recs = []
    try:
        for addr in ec2.describe_addresses().get('Addresses', []):
            alloc = addr.get('AllocationId', addr.get('PublicIp', ''))
            name  = next((t['Value'] for t in addr.get('Tags', [])
                          if t['Key'] == 'Name'), addr.get('PublicIp', alloc))
            arn   = f"arn:aws:ec2:{region}:{account_id}:elastic-ip/{alloc}"
            assoc = addr.get('AssociationId', '')
            recs.append(make_rec(
                account_id, name, 'ec2:elasticIp', 'ec2', arn, region,
                tags=addr.get('Tags', []),
                status='associated' if assoc else 'unassociated',
                unassigned='No' if assoc else 'Yes',
            ))
    except Exception:
        pass
    return recs


def scan_ec2_nat_gateways(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_nat_gateways').paginate():
            for ng in page['NatGateways']:
                ngid = ng['NatGatewayId']
                name = next((t['Value'] for t in ng.get('Tags', [])
                             if t['Key'] == 'Name'), ngid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:natgateway/{ngid}"
                recs.append(make_rec(
                    account_id, name, 'ec2:natGateway', 'ec2', arn, region,
                    tags=ng.get('Tags', []),
                    created=fmt_date(ng.get('CreateTime')),
                    status=ng.get('State', ''),
                ))
    except Exception:
        pass
    return recs


def scan_ec2_internet_gateways(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_internet_gateways').paginate():
            for ig in page['InternetGateways']:
                igid = ig['InternetGatewayId']
                name = next((t['Value'] for t in ig.get('Tags', [])
                             if t['Key'] == 'Name'), igid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:internet-gateway/{igid}"
                atts = ig.get('Attachments', [])
                recs.append(make_rec(
                    account_id, name, 'ec2:internetGateway', 'ec2', arn, region,
                    tags=ig.get('Tags', []),
                    status='attached' if atts else 'detached',
                    unassigned='No' if atts else 'Yes',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_network_interfaces(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_network_interfaces').paginate():
            for ni in page['NetworkInterfaces']:
                niid = ni['NetworkInterfaceId']
                name = ni.get('Description', niid) or niid
                arn  = f"arn:aws:ec2:{region}:{account_id}:network-interface/{niid}"
                att  = ni.get('Attachment')
                recs.append(make_rec(
                    account_id, name, 'ec2:networkInterface', 'ec2', arn, region,
                    tags=ni.get('TagSet', []),
                    status=ni.get('Status', ''),
                    unassigned='No' if att else 'Yes',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_key_pairs(ec2, account_id, region):
    recs = []
    try:
        for kp in ec2.describe_key_pairs().get('KeyPairs', []):
            name = kp['KeyName']
            arn  = f"arn:aws:ec2:{region}:{account_id}:key-pair/{kp.get('KeyPairId', name)}"
            recs.append(make_rec(
                account_id, name, 'ec2:keyPair', 'ec2', arn, region,
                tags=kp.get('Tags', []),
                created=fmt_date(kp.get('CreateTime')),
                status='active',
            ))
    except Exception:
        pass
    return recs


def scan_ec2_amis(ec2, account_id, region):
    recs = []
    try:
        for img in ec2.describe_images(Owners=[account_id]).get('Images', []):
            iid  = img['ImageId']
            arn  = f"arn:aws:ec2:{region}:{account_id}:image/{iid}"
            recs.append(make_rec(
                account_id, img.get('Name', iid), 'ec2:image', 'ec2', arn, region,
                tags=img.get('Tags', []),
                created=img.get('CreationDate', '')[:10],
                status=img.get('State', ''),
            ))
    except Exception:
        pass
    return recs


def scan_ec2_launch_templates(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_launch_templates').paginate():
            for lt in page['LaunchTemplates']:
                ltid = lt['LaunchTemplateId']
                arn  = f"arn:aws:ec2:{region}:{account_id}:launch-template/{ltid}"
                recs.append(make_rec(
                    account_id, lt.get('LaunchTemplateName', ltid),
                    'ec2:launchTemplate', 'ec2', arn, region,
                    tags=lt.get('Tags', []),
                    created=fmt_date(lt.get('CreateTime')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_ec2_transit_gateways(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_transit_gateways').paginate():
            for tg in page['TransitGateways']:
                name = next((t['Value'] for t in tg.get('Tags', [])
                             if t['Key'] == 'Name'), tg['TransitGatewayId'])
                recs.append(make_rec(
                    account_id, name, 'ec2:transitGateway', 'ec2',
                    tg['TransitGatewayArn'], region,
                    tags=tg.get('Tags', []),
                    created=fmt_date(tg.get('CreationTime')),
                    status=tg.get('State', ''),
                ))
    except Exception:
        pass
    return recs


def scan_ec2_route_tables(ec2, account_id, region):
    recs = []
    try:
        for page in ec2.get_paginator('describe_route_tables').paginate():
            for rt in page['RouteTables']:
                rtid = rt['RouteTableId']
                name = next((t['Value'] for t in rt.get('Tags', [])
                             if t['Key'] == 'Name'), rtid)
                arn  = f"arn:aws:ec2:{region}:{account_id}:route-table/{rtid}"
                recs.append(make_rec(
                    account_id, name, 'ec2:routeTable', 'ec2', arn, region,
                    tags=rt.get('Tags', []),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def _get_rds_tags(rds, arn):
    try:
        return rds.list_tags_for_resource(ResourceName=arn).get('TagList', [])
    except Exception:
        return []


def scan_rds(rds, account_id, region):
    recs = []
    try:
        for page in rds.get_paginator('describe_db_instances').paginate():
            for db in page['DBInstances']:
                arn  = db['DBInstanceArn']
                tags = _get_rds_tags(rds, arn)
                recs.append(make_rec(
                    account_id, db['DBInstanceIdentifier'],
                    'rds:db', 'rds', arn, region,
                    tags=tags,
                    created=fmt_date(db.get('InstanceCreateTime')),
                    status=db.get('DBInstanceStatus', ''),
                ))
        for page in rds.get_paginator('describe_db_clusters').paginate():
            for cl in page['DBClusters']:
                arn  = cl['DBClusterArn']
                tags = _get_rds_tags(rds, arn)
                recs.append(make_rec(
                    account_id, cl['DBClusterIdentifier'],
                    'rds:cluster', 'rds', arn, region,
                    tags=tags,
                    created=fmt_date(cl.get('ClusterCreateTime')),
                    status=cl.get('Status', ''),
                ))
    except Exception:
        pass
    return recs


def scan_s3(s3, account_id):
    recs = []
    try:
        for b in s3.list_buckets().get('Buckets', []):
            name = b['Name']
            arn  = f"arn:aws:s3:::{name}"
            try:
                loc     = s3.get_bucket_location(Bucket=name)
                bregion = loc.get('LocationConstraint') or 'us-east-1'
            except Exception:
                bregion = 'us-east-1'
            try:
                tags = s3.get_bucket_tagging(Bucket=name).get('TagSet', [])
            except Exception:
                tags = []
            recs.append(make_rec(
                account_id, name, 's3:bucket', 's3', arn, bregion,
                tags=tags,
                created=fmt_date(b.get('CreationDate')),
                status='active',
            ))
    except Exception:
        pass
    return recs


def scan_lambda(lmb, account_id, region):
    recs = []
    try:
        for page in lmb.get_paginator('list_functions').paginate():
            for fn in page.get('Functions', []):
                recs.append(make_rec(
                    account_id, fn['FunctionName'],
                    'lambda:function', 'lambda',
                    fn['FunctionArn'], region,
                    created=fn.get('LastModified', '')[:10],
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_iam_users(iam, account_id):
    recs = []
    try:
        for page in iam.get_paginator('list_users').paginate():
            for u in page['Users']:
                last = u.get('PasswordLastUsed')
                if last:
                    days   = (datetime.datetime.utcnow().replace(tzinfo=None)
                              - last.replace(tzinfo=None)).days
                    status = 'active' if days < 90 else 'inactive (>90d)'
                else:
                    status = 'inactive (never logged in)'
                recs.append(make_rec(
                    account_id, u['UserName'],
                    'iam:user', 'iam', u['Arn'], 'global',
                    created=fmt_date(u.get('CreateDate')),
                    status=status,
                ))
    except Exception:
        pass
    return recs


def scan_iam_roles(iam, account_id):
    recs = []
    try:
        for page in iam.get_paginator('list_roles').paginate():
            for r in page['Roles']:
                recs.append(make_rec(
                    account_id, r['RoleName'],
                    'iam:role', 'iam', r['Arn'], 'global',
                    created=fmt_date(r.get('CreateDate')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_iam_policies(iam, account_id):
    recs = []
    try:
        for page in iam.get_paginator('list_policies').paginate(Scope='Local'):
            for p in page['Policies']:
                att = p.get('AttachmentCount', 0)
                recs.append(make_rec(
                    account_id, p['PolicyName'],
                    'iam:policy', 'iam', p['Arn'], 'global',
                    created=fmt_date(p.get('CreateDate')),
                    status='attached' if att > 0 else 'unattached',
                    unassigned='No' if att > 0 else 'Yes',
                ))
    except Exception:
        pass
    return recs


def scan_iam_groups(iam, account_id):
    recs = []
    try:
        for page in iam.get_paginator('list_groups').paginate():
            for g in page['Groups']:
                recs.append(make_rec(
                    account_id, g['GroupName'],
                    'iam:group', 'iam', g['Arn'], 'global',
                    created=fmt_date(g.get('CreateDate')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_dynamodb(ddb, account_id, region):
    recs = []
    try:
        for page in ddb.get_paginator('list_tables').paginate():
            for tname in page.get('TableNames', []):
                try:
                    td = ddb.describe_table(TableName=tname)['Table']
                    recs.append(make_rec(
                        account_id, tname,
                        'dynamodb:table', 'dynamodb',
                        td['TableArn'], region,
                        created=fmt_date(td.get('CreationDateTime')),
                        status=td.get('TableStatus', ''),
                    ))
                except Exception:
                    pass
    except Exception:
        pass
    return recs


def scan_elb(elb, account_id, region):
    recs = []
    try:
        for page in elb.get_paginator('describe_load_balancers').paginate():
            for lb in page['LoadBalancers']:
                recs.append(make_rec(
                    account_id, lb['LoadBalancerName'],
                    'elasticloadbalancing:loadBalancer',
                    'elasticloadbalancing',
                    lb['LoadBalancerArn'], region,
                    created=fmt_date(lb.get('CreatedTime')),
                    status=lb.get('State', {}).get('Code', ''),
                ))
    except Exception:
        pass
    return recs


def scan_cloudformation(cf, account_id, region):
    recs = []
    try:
        for page in cf.get_paginator('describe_stacks').paginate():
            for stack in page.get('Stacks', []):
                recs.append(make_rec(
                    account_id, stack['StackName'],
                    'cloudformation:stack', 'cloudformation',
                    stack['StackId'], region,
                    tags=stack.get('Tags', []),
                    created=fmt_date(stack.get('CreationTime')),
                    status=stack.get('StackStatus', ''),
                ))
    except Exception:
        pass
    return recs


def scan_secretsmanager(sm, account_id, region):
    recs = []
    try:
        for page in sm.get_paginator('list_secrets').paginate():
            for sec in page.get('SecretList', []):
                recs.append(make_rec(
                    account_id, sec['Name'],
                    'secretsmanager:secret', 'secretsmanager',
                    sec['ARN'], region,
                    tags=sec.get('Tags', []),
                    created=fmt_date(sec.get('CreatedDate')),
                    status='deleted' if sec.get('DeletedDate') else 'active',
                ))
    except Exception:
        pass
    return recs


def scan_sns(sns, account_id, region):
    recs = []
    try:
        for page in sns.get_paginator('list_topics').paginate():
            for t in page.get('Topics', []):
                arn  = t['TopicArn']
                name = arn.split(':')[-1]
                recs.append(make_rec(
                    account_id, name, 'sns:topic', 'sns', arn, region,
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_sqs(sqs, account_id, region):
    recs = []
    try:
        for page in sqs.get_paginator('list_queues').paginate():
            for url in page.get('QueueUrls', []):
                name = url.split('/')[-1]
                arn  = f"arn:aws:sqs:{region}:{account_id}:{name}"
                recs.append(make_rec(
                    account_id, name, 'sqs:queue', 'sqs', arn, region,
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_kms(kms, account_id, region):
    recs = []
    try:
        for page in kms.get_paginator('list_keys').paginate():
            for key in page.get('Keys', []):
                kid = key['KeyId']
                try:
                    meta = kms.describe_key(KeyId=kid)['KeyMetadata']
                    if meta.get('KeyManager') == 'AWS':
                        continue
                    recs.append(make_rec(
                        account_id, kid, 'kms:key', 'kms',
                        meta['Arn'], region,
                        created=fmt_date(meta.get('CreationDate')),
                        status=meta.get('KeyState', ''),
                    ))
                except Exception:
                    pass
    except Exception:
        pass
    return recs


def scan_eks(eks, account_id, region):
    recs = []
    try:
        for page in eks.get_paginator('list_clusters').paginate():
            for cname in page.get('clusters', []):
                try:
                    cl = eks.describe_cluster(name=cname)['cluster']
                    recs.append(make_rec(
                        account_id, cname, 'eks:cluster', 'eks',
                        cl['arn'], region,
                        tags=cl.get('tags', {}),
                        created=fmt_date(cl.get('createdAt')),
                        status=cl.get('status', ''),
                    ))
                except Exception:
                    pass
    except Exception:
        pass
    return recs


def scan_ecs_clusters(ecs, account_id, region):
    recs = []
    try:
        arns = []
        for page in ecs.get_paginator('list_clusters').paginate():
            arns.extend(page.get('clusterArns', []))
        if arns:
            for cl in ecs.describe_clusters(clusters=arns).get('clusters', []):
                recs.append(make_rec(
                    account_id, cl['clusterName'],
                    'ecs:cluster', 'ecs', cl['clusterArn'], region,
                    status=cl.get('status', ''),
                ))
    except Exception:
        pass
    return recs


def scan_elasticache(ec, account_id, region):
    recs = []
    try:
        for page in ec.get_paginator('describe_cache_clusters').paginate():
            for cl in page.get('CacheClusters', []):
                recs.append(make_rec(
                    account_id, cl['CacheClusterId'],
                    'elasticache:cluster', 'elasticache',
                    cl.get('ARN', ''), region,
                    created=fmt_date(cl.get('CacheClusterCreateTime')),
                    status=cl.get('CacheClusterStatus', ''),
                ))
    except Exception:
        pass
    return recs


def scan_cloudwatch_alarms(cw, account_id, region):
    recs = []
    try:
        for page in cw.get_paginator('describe_alarms').paginate():
            for alarm in page.get('MetricAlarms', []):
                recs.append(make_rec(
                    account_id, alarm['AlarmName'],
                    'cloudwatch:alarm', 'cloudwatch',
                    alarm['AlarmArn'], region,
                    status=alarm.get('StateValue', ''),
                ))
    except Exception:
        pass
    return recs


def scan_logs(logs, account_id, region):
    recs = []
    try:
        for page in logs.get_paginator('describe_log_groups').paginate():
            for lg in page.get('logGroups', []):
                arn = lg.get('arn', '').rstrip(':*')
                ms  = lg.get('creationTime', 0)
                created = (datetime.datetime.utcfromtimestamp(ms / 1000)
                           .strftime('%Y-%m-%d') if ms else '')
                recs.append(make_rec(
                    account_id, lg['logGroupName'],
                    'logs:logGroup', 'logs', arn, region,
                    created=created,
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_stepfunctions(sfn, account_id, region):
    recs = []
    try:
        for page in sfn.get_paginator('list_state_machines').paginate():
            for sm in page.get('stateMachines', []):
                recs.append(make_rec(
                    account_id, sm['name'],
                    'stepfunctions:stateMachine', 'stepfunctions',
                    sm['stateMachineArn'], region,
                    created=fmt_date(sm.get('creationDate')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_ecr(ecr, account_id, region):
    recs = []
    try:
        for page in ecr.get_paginator('describe_repositories').paginate():
            for repo in page.get('repositories', []):
                recs.append(make_rec(
                    account_id, repo['repositoryName'],
                    'ecr:repository', 'ecr',
                    repo['repositoryArn'], region,
                    created=fmt_date(repo.get('createdAt')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_acm(acm, account_id, region):
    recs = []
    try:
        for page in acm.get_paginator('list_certificates').paginate():
            for cert in page.get('CertificateSummaryList', []):
                arn  = cert['CertificateArn']
                name = cert.get('DomainName', arn.split('/')[-1])
                recs.append(make_rec(
                    account_id, name, 'acm:certificate', 'acm', arn, region,
                    status=cert.get('Status', ''),
                ))
    except Exception:
        pass
    return recs


def scan_redshift(rs, account_id, region):
    recs = []
    try:
        for page in rs.get_paginator('describe_clusters').paginate():
            for cl in page.get('Clusters', []):
                cid = cl['ClusterIdentifier']
                arn = f"arn:aws:redshift:{region}:{account_id}:cluster:{cid}"
                recs.append(make_rec(
                    account_id, cid, 'redshift:cluster', 'redshift', arn, region,
                    created=fmt_date(cl.get('ClusterCreateTime')),
                    status=cl.get('ClusterStatus', ''),
                ))
    except Exception:
        pass
    return recs


def scan_kinesis(kin, account_id, region):
    recs = []
    try:
        for page in kin.get_paginator('list_streams').paginate():
            for name in page.get('StreamNames', []):
                try:
                    sd = kin.describe_stream_summary(StreamName=name)['StreamDescriptionSummary']
                    recs.append(make_rec(
                        account_id, name, 'kinesis:stream', 'kinesis',
                        sd['StreamARN'], region,
                        created=fmt_date(sd.get('StreamCreationTimestamp')),
                        status=sd.get('StreamStatus', ''),
                    ))
                except Exception:
                    pass
    except Exception:
        pass
    return recs


def scan_ssm_parameters(ssm, account_id, region):
    recs = []
    try:
        for page in ssm.get_paginator('describe_parameters').paginate():
            for p in page.get('Parameters', []):
                name = p['Name']
                arn  = f"arn:aws:ssm:{region}:{account_id}:parameter{name}"
                recs.append(make_rec(
                    account_id, name, 'ssm:parameter', 'ssm', arn, region,
                    created=fmt_date(p.get('LastModifiedDate')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_route53(r53, account_id):
    recs = []
    try:
        for page in r53.get_paginator('list_hosted_zones').paginate():
            for hz in page.get('HostedZones', []):
                hzid = hz['Id'].split('/')[-1]
                arn  = f"arn:aws:route53:::hostedzone/{hzid}"
                recs.append(make_rec(
                    account_id, hz['Name'],
                    'route53:hostedZone', 'route53', arn, 'global',
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_cloudfront(cf, account_id):
    recs = []
    try:
        for page in cf.get_paginator('list_distributions').paginate():
            for dist in page.get('DistributionList', {}).get('Items', []):
                recs.append(make_rec(
                    account_id, dist.get('DomainName', dist['Id']),
                    'cloudfront:distribution', 'cloudfront',
                    dist['ARN'], 'global',
                    status=dist.get('Status', ''),
                ))
    except Exception:
        pass
    return recs


def scan_glue_jobs(glue, account_id, region):
    recs = []
    try:
        for page in glue.get_paginator('get_jobs').paginate():
            for job in page.get('Jobs', []):
                name = job['Name']
                arn  = f"arn:aws:glue:{region}:{account_id}:job/{name}"
                recs.append(make_rec(
                    account_id, name, 'glue:job', 'glue', arn, region,
                    created=fmt_date(job.get('CreatedOn')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_codepipeline(cp, account_id, region):
    recs = []
    try:
        for page in cp.get_paginator('list_pipelines').paginate():
            for p in page.get('pipelines', []):
                name = p['name']
                arn  = f"arn:aws:codepipeline:{region}:{account_id}:{name}"
                recs.append(make_rec(
                    account_id, name,
                    'codepipeline:pipeline', 'codepipeline', arn, region,
                    created=fmt_date(p.get('created')),
                    status='active',
                ))
    except Exception:
        pass
    return recs


def scan_codebuild(cb, account_id, region):
    recs = []
    try:
        for page in cb.get_paginator('list_projects').paginate():
            names = page.get('projects', [])
            if names:
                for proj in cb.batch_get_projects(names=names).get('projects', []):
                    recs.append(make_rec(
                        account_id, proj['name'],
                        'codebuild:project', 'codebuild',
                        proj['arn'], region,
                        created=fmt_date(proj.get('created')),
                        status='active',
                    ))
    except Exception:
        pass
    return recs


def scan_sagemaker(sm, account_id, region):
    recs = []
    try:
        for page in sm.get_paginator('list_endpoints').paginate():
            for ep in page.get('Endpoints', []):
                recs.append(make_rec(
                    account_id, ep['EndpointName'],
                    'sagemaker:endpoint', 'sagemaker',
                    ep['EndpointArn'], region,
                    created=fmt_date(ep.get('CreationTime')),
                    status=ep.get('EndpointStatus', ''),
                ))
    except Exception:
        pass
    return recs


def scan_wafv2(waf, account_id, region):
    recs = []
    try:
        for page in waf.get_paginator('list_web_acls').paginate(Scope='REGIONAL'):
            for acl in page.get('WebACLs', []):
                recs.append(make_rec(
                    account_id, acl['Name'],
                    'wafv2:webAcl', 'wafv2',
                    acl['ARN'], region,
                    status='active',
                ))
    except Exception:
        pass
    return recs


# ══════════════════════════════════════════════════════════════════════════════
# CLOUDTRAIL  — fill Created By
# ══════════════════════════════════════════════════════════════════════════════

def _who(ev):
    who = ev.get('Username', '')
    if not who:
        try:
            parsed = json.loads(ev.get('CloudTrailEvent', '{}'))
            uid    = parsed.get('userIdentity', {})
            who    = (
                uid.get('userName')
                or uid.get('sessionContext', {}).get('sessionIssuer', {}).get('userName')
                or uid.get('arn', '').split('/')[-1]
                or uid.get('principalId', '').split(':')[-1]
                or uid.get('type', '')
            )
        except Exception:
            pass
    return who or 'Unknown'


def build_cloudtrail_cache(regions, days_back=90):
    print(f"\n  Building CloudTrail cache (last {days_back} days, {len(regions)} regions)...")
    cache      = {}
    end_time   = datetime.datetime.utcnow()
    start_time = end_time - datetime.timedelta(days=days_back)

    for region in regions:
        try:
            ct = boto3.client('cloudtrail', region_name=region)
            for event_name in CREATE_EVENTS:
                token = None
                while True:
                    kwargs = {
                        'LookupAttributes': [
                            {'AttributeKey': 'EventName', 'AttributeValue': event_name}
                        ],
                        'StartTime'  : start_time,
                        'EndTime'    : end_time,
                        'MaxResults' : 50,
                    }
                    if token:
                        kwargs['NextToken'] = token
                    try:
                        resp = ct.lookup_events(**kwargs)
                    except Exception:
                        break
                    for ev in resp.get('Events', []):
                        who = _who(ev)
                        for res in ev.get('Resources', []):
                            rname = res.get('ResourceName', '')
                            if rname and rname not in cache:
                                cache[rname] = who
                    token = resp.get('NextToken')
                    if not token:
                        break
        except Exception:
            continue

    print(f"  CloudTrail cache: {len(cache)} entries")
    return cache


def apply_cloudtrail(records, cache):
    for rec in records:
        arn  = rec.get('ARN', '')
        name = rec.get('Resource Name', '')
        # try full ARN, then ARN suffix, then resource name
        suffix = arn.split('/')[-1] if '/' in arn else arn.split(':')[-1]
        who = cache.get(arn) or cache.get(suffix) or cache.get(name) or 'Unknown'
        rec['Created By'] = who


# ══════════════════════════════════════════════════════════════════════════════
# INTERRUPT HANDLER
# ══════════════════════════════════════════════════════════════════════════════

_INTERRUPT_RECORDS = []
_INTERRUPT_ACCOUNT = ''


def _handle_interrupt(signum, frame):
    print("\n\n  *** Scan interrupted ***")
    if _INTERRUPT_RECORDS:
        print(f"  Saving partial Excel ({len(_INTERRUPT_RECORDS)} resources)...")
        write_excel(_INTERRUPT_RECORDS, _INTERRUPT_ACCOUNT, partial=True)
    sys.exit(0)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN SCAN
# ══════════════════════════════════════════════════════════════════════════════

def scan_all(account_id, regions):
    global _INTERRUPT_RECORDS, _INTERRUPT_ACCOUNT
    _INTERRUPT_ACCOUNT = account_id
    records   = []
    seen_arns = set()

    def add(recs):
        for r in recs:
            arn = r.get('ARN', '')
            if arn and arn in seen_arns:
                continue
            if arn:
                seen_arns.add(arn)
            records.append(r)
            _INTERRUPT_RECORDS = records

    # Global services
    print("\n  [Global] IAM...")
    iam = boto3.client('iam', region_name='us-east-1')
    add(scan_iam_users(iam, account_id))
    add(scan_iam_roles(iam, account_id))
    add(scan_iam_policies(iam, account_id))
    add(scan_iam_groups(iam, account_id))

    print("  [Global] S3...")
    add(scan_s3(boto3.client('s3', region_name='us-east-1'), account_id))

    print("  [Global] Route53...")
    add(scan_route53(boto3.client('route53', region_name='us-east-1'), account_id))

    print("  [Global] CloudFront...")
    add(scan_cloudfront(boto3.client('cloudfront', region_name='us-east-1'), account_id))

    # Regional services
    total = len(regions)
    for idx, region in enumerate(regions, 1):
        print(f"\n  [{idx:>2}/{total}] {region}")

        def c(svc, r=region):
            return boto3.client(svc, region_name=r)

        ec2 = c('ec2')
        for fn in [
            scan_ec2_instances, scan_ec2_volumes, scan_ec2_snapshots,
            scan_ec2_security_groups, scan_ec2_vpcs, scan_ec2_subnets,
            scan_ec2_eips, scan_ec2_nat_gateways, scan_ec2_internet_gateways,
            scan_ec2_network_interfaces, scan_ec2_key_pairs, scan_ec2_amis,
            scan_ec2_launch_templates, scan_ec2_transit_gateways,
            scan_ec2_route_tables,
        ]:
            recs = fn(ec2, account_id, region)
            if recs:
                print(f"    {fn.__name__.replace('scan_ec2_','ec2:')} → {len(recs)}")
            add(recs)

        SVCS = [
            ('rds',             scan_rds,              'rds'),
            ('lambda',          scan_lambda,           'lambda'),
            ('dynamodb',        scan_dynamodb,         'dynamodb'),
            ('elbv2',           scan_elb,              'elb'),
            ('cloudformation',  scan_cloudformation,   'cloudformation'),
            ('secretsmanager',  scan_secretsmanager,   'secretsmanager'),
            ('sns',             scan_sns,              'sns'),
            ('sqs',             scan_sqs,              'sqs'),
            ('kms',             scan_kms,              'kms'),
            ('eks',             scan_eks,              'eks'),
            ('ecs',             scan_ecs_clusters,     'ecs'),
            ('elasticache',     scan_elasticache,      'elasticache'),
            ('cloudwatch',      scan_cloudwatch_alarms,'cloudwatch'),
            ('logs',            scan_logs,             'logs'),
            ('stepfunctions',   scan_stepfunctions,    'stepfunctions'),
            ('ecr',             scan_ecr,              'ecr'),
            ('acm',             scan_acm,              'acm'),
            ('redshift',        scan_redshift,         'redshift'),
            ('kinesis',         scan_kinesis,          'kinesis'),
            ('ssm',             scan_ssm_parameters,   'ssm'),
            ('glue',            scan_glue_jobs,        'glue'),
            ('codepipeline',    scan_codepipeline,     'codepipeline'),
            ('codebuild',       scan_codebuild,        'codebuild'),
            ('sagemaker',       scan_sagemaker,        'sagemaker'),
            ('wafv2',           scan_wafv2,            'wafv2'),
        ]

        for boto_svc, fn, label in SVCS:
            try:
                recs = fn(c(boto_svc), account_id, region)
                if recs:
                    print(f"    {label} → {len(recs)}")
                add(recs)
            except Exception:
                pass

        print(f"    ✓ total so far: {len(records)}")

    print(f"\n  Total unique resources: {len(records)}")
    return records


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_summary_sheet(wb, records, account_id, partial=False):
    ws = wb.create_sheet(title='Summary', index=0)
    ws.merge_cells('A1:E1')
    ws['A1'].value     = (f'AWS Inventory — {account_id}'
                          + ('  ⚠ PARTIAL' if partial else ''))
    ws['A1'].font      = TITLE_FONT
    ws['A1'].fill      = HEADER_FILL
    ws['A1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:E2')
    ts = NOW_UTC.strftime('%Y-%m-%d %H:%M UTC')
    ws['A2'].value     = f'Generated: {ts}   |   Total: {len(records)}'
    ws['A2'].font      = Font(italic=True, name='Calibri', size=9, color='555555')
    ws['A2'].alignment = CENTER_ALIGN
    ws.row_dimensions[2].height = 18

    for ci, h in enumerate(['Service','Resource Type','Count','Active','Inactive/Unassigned'], 1):
        cell = ws.cell(4, ci, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.border = BORDER;    cell.alignment = CENTER_ALIGN
    ws.row_dimensions[4].height = 24

    agg = defaultdict(lambda: defaultdict(lambda: {'count':0,'active':0,'inactive':0}))
    for rec in records:
        svc    = rec['Service']
        rtype  = rec['Resource Type']
        status = str(rec.get('Status', '')).lower()
        unasgn = str(rec.get('Unassigned', '')).lower()
        agg[svc][rtype]['count'] += 1
        if any(w in status for w in ACTIVE_WORDS):
            agg[svc][rtype]['active'] += 1
        elif any(w in status for w in INACTIVE_WORDS) or unasgn == 'yes':
            agg[svc][rtype]['inactive'] += 1

    ri = 5
    for fi, svc in enumerate(sorted(agg)):
        fill = SVC_FILLS[fi % len(SVC_FILLS)]
        for rtype in sorted(agg[svc]):
            info = agg[svc][rtype]
            for ci, val in enumerate([svc, rtype,
                                       info['count'], info['active'], info['inactive']], 1):
                cell = ws.cell(ri, ci, val)
                cell.font = BODY_FONT; cell.border = BORDER; cell.fill = fill
                cell.alignment = CENTER_ALIGN if ci >= 3 else LEFT_ALIGN
            ri += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 22
    ws.freeze_panes    = 'A5'
    ws.auto_filter.ref = f'A4:E{ri-1}'


def write_data_sheet(wb, title, records):
    ws  = wb.create_sheet(title=title[:31])
    sci = COLS.index('Status')     + 1
    uci = COLS.index('Unassigned') + 1

    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(1, ci, col)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.border = BORDER;    cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for ri, rec in enumerate(records, 2):
        bg = ALT_FILL if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(COLS, 1):
            val  = rec.get(col, '') or ''
            cell = ws.cell(ri, ci, val)
            cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = LEFT_ALIGN
            if bg: cell.fill = bg

        status = str(rec.get('Status', '')).lower()
        sc = ws.cell(ri, sci)
        if any(w in status for w in ACTIVE_WORDS):
            sc.fill = GREEN_FILL
            sc.font = Font(name='Calibri', size=10, color='1B5E20')
        elif any(w in status for w in INACTIVE_WORDS):
            sc.fill = RED_FILL
            sc.font = Font(name='Calibri', size=10, color='B71C1C')
        elif status:
            sc.fill = ORANGE_FILL
            sc.font = Font(name='Calibri', size=10, color='E65100')

        uval = str(rec.get('Unassigned', '')).lower()
        uc   = ws.cell(ri, uci)
        if uval == 'yes':
            uc.fill = ORANGE_FILL
        elif uval == 'no':
            uc.fill = GREEN_FILL

    cw = {ci: len(COLS[ci-1]) for ci in range(1, len(COLS)+1)}
    for rec in records:
        for ci, col in enumerate(COLS, 1):
            cw[ci] = max(cw[ci], min(len(str(rec.get(col,'') or '')), 50))
    for ci, w in cw.items():
        ws.column_dimensions[get_column_letter(ci)].width = w + 3


def write_excel(records, account_id, partial=False):
    ts       = NOW_UTC.strftime('%Y%m%d_%H%M')
    suffix   = '_PARTIAL' if partial else ''
    filename = f"aws_inventory_{account_id}_{ts}{suffix}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, records, account_id, partial=partial)
    write_data_sheet(wb, 'All Resources', records)

    by_svc = defaultdict(list)
    for rec in records:
        by_svc[rec['Service']].append(rec)

    for svc in sorted(by_svc):
        write_data_sheet(wb, svc[:31], by_svc[svc])
        print(f"  Sheet: {svc:<30} ({len(by_svc[svc])})")

    wb.save(filename)
    print(f"\n  Saved   : {filename}")
    print(f"  Services: {len(by_svc)}")
    print(f"  Total   : {len(records)}")
    return filename


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    signal.signal(signal.SIGINT,  _handle_interrupt)
    signal.signal(signal.SIGTERM, _handle_interrupt)

    print("=" * 55)
    print("  AWS All-Resource Inventory Scanner")
    print("=" * 55)

    print("\n[1/4] Resolving account...")
    account_id = get_account_id()
    print(f"  Account ID: {account_id}")

    print("\n[2/4] Discovering regions...")
    regions = get_all_regions()
    print(f"  Regions ({len(regions)}): {', '.join(regions)}")

    print("\n[3/4] Scanning all resources...")
    print("       (Ctrl+C at any time — partial Excel will be saved)\n")
    records = scan_all(account_id, regions)

    if not records:
        print("  No resources found.")
        sys.exit(0)

    print("\n[4/4] CloudTrail creator lookup + writing Excel...")
    cache = build_cloudtrail_cache(regions, days_back=90)
    apply_cloudtrail(records, cache)
    write_excel(records, account_id, partial=False)
    print("\n  Done.")
